"""Per-user business document management."""
import json
import sqlite3
from pathlib import Path

DB_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "serviceflow.db"
UPLOADS_ROOT = Path(__file__).resolve().parent.parent.parent / "data" / "uploads" / "users"

# Canonical filename → human-readable label
DOCUMENT_TYPES: dict[str, str] = {
    "company_profile.json":       "Company Profile",
    "service_catalogue.json":     "Service Catalogue",
    "pricing_data.json":          "Pricing Data",
    "business_policies.txt":      "Business Policies",
    "staffing_availability.json": "Staffing Availability",
    "schedule_capacity.json":     "Schedule Capacity",
    "service_area.json":          "Service Area",
    "equipment_readiness.json":   "Equipment Readiness",
    "risk_flags.json":            "Risk Flags",
    "job_complexity.json":        "Job Complexity",
}

# Reverse: human-readable label → canonical filename
LABEL_TO_KEY: dict[str, str] = {v: k for k, v in DOCUMENT_TYPES.items()}

_DDL = """
CREATE TABLE IF NOT EXISTS business_documents (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id     INTEGER NOT NULL REFERENCES users(id),
    doc_type    TEXT NOT NULL,
    filename    TEXT NOT NULL,
    stored_path TEXT NOT NULL,
    uploaded_at TEXT DEFAULT (datetime('now')),
    UNIQUE (user_id, doc_type)
)
"""

_DECISIONS_DDL = """
CREATE TABLE IF NOT EXISTS quote_decisions (
    id         INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id    INTEGER NOT NULL REFERENCES users(id),
    decision   TEXT NOT NULL CHECK(decision IN ('approved', 'rejected')),
    created_at TEXT DEFAULT (datetime('now'))
)
"""


def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def ensure_table() -> None:
    with _conn() as conn:
        conn.execute(_DDL)
        conn.execute(_DECISIONS_DDL)


def record_decision(user_id: int, decision: str) -> None:
    with _conn() as conn:
        conn.execute(
            "INSERT INTO quote_decisions (user_id, decision) VALUES (?, ?)",
            (user_id, decision),
        )


def get_decision_counts(user_id: int) -> dict[str, int]:
    with _conn() as conn:
        rows = conn.execute(
            "SELECT decision, COUNT(*) AS cnt FROM quote_decisions "
            "WHERE user_id = ? GROUP BY decision",
            (user_id,),
        ).fetchall()
    counts: dict[str, int] = {"approved": 0, "rejected": 0}
    for row in rows:
        counts[row["decision"]] = row["cnt"]
    return counts


def get_user_dir(user_id: int) -> Path:
    d = UPLOADS_ROOT / str(user_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _parse_to_canonical(tmp_path: str, canonical_name: str) -> str:
    """Parse any supported file type into the canonical storage format (JSON or plain text)."""
    import pdfplumber
    import docx as _docx
    import openpyxl

    path = Path(tmp_path)
    ext = path.suffix.lower()
    canonical_ext = Path(canonical_name).suffix.lower()

    if ext == ".pdf":
        pages = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    pages.append(text.strip())
        raw_text = "\n\n".join(pages)

    elif ext in (".docx", ".doc"):
        doc = _docx.Document(path)
        raw_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    elif ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets: dict = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [
                [str(c) if c is not None else "" for c in row]
                for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)
            ]
            if rows:
                sheets[name] = rows
        wb.close()
        raw_text = json.dumps(sheets, ensure_ascii=False)

    elif ext == ".json":
        raw_text = path.read_text(encoding="utf-8")

    else:
        raw_text = path.read_text(encoding="utf-8", errors="replace")

    if canonical_ext == ".json":
        try:
            json.loads(raw_text)
            return raw_text
        except (json.JSONDecodeError, ValueError):
            return json.dumps({"content": raw_text}, ensure_ascii=False, indent=2)
    else:
        return raw_text


def upload_document(user_id: int, doc_type: str, tmp_path: str) -> tuple[bool, str]:
    """Parse the uploaded file into the canonical format and record it in the DB."""
    if doc_type not in DOCUMENT_TYPES:
        return False, f"Unknown document type: {doc_type}"

    dest = get_user_dir(user_id) / doc_type
    try:
        content = _parse_to_canonical(tmp_path, doc_type)
        dest.write_text(content, encoding="utf-8")
    except Exception as exc:
        return False, f"Could not process file: {exc}"

    try:
        with _conn() as conn:
            conn.execute(
                """INSERT INTO business_documents (user_id, doc_type, filename, stored_path)
                   VALUES (?, ?, ?, ?)
                   ON CONFLICT(user_id, doc_type) DO UPDATE
                   SET filename=excluded.filename,
                       stored_path=excluded.stored_path,
                       uploaded_at=datetime('now')""",
                (user_id, doc_type, Path(tmp_path).name, str(dest)),
            )
        return True, f"{DOCUMENT_TYPES[doc_type]} uploaded successfully."
    except Exception as exc:
        return False, f"DB error: {exc}"


def get_user_documents(user_id: int) -> list[dict]:
    with _conn() as conn:
        rows = conn.execute(
            "SELECT doc_type, uploaded_at FROM business_documents WHERE user_id = ? ORDER BY uploaded_at DESC",
            (user_id,),
        ).fetchall()
    return [{"doc_type": r["doc_type"], "label": DOCUMENT_TYPES.get(r["doc_type"], r["doc_type"]),
             "uploaded_at": r["uploaded_at"]} for r in rows]


def has_documents(user_id: int) -> bool:
    with _conn() as conn:
        row = conn.execute(
            "SELECT 1 FROM business_documents WHERE user_id = ? LIMIT 1", (user_id,)
        ).fetchone()
    return row is not None


def build_docs_table_html(user_id: int, page: int = 0, per_page: int = 5) -> str:
    docs = get_user_documents(user_id)
    if not docs:
        return (
            '<div style="background:#fef3c7;border:1px solid #fcd34d;border-radius:8px;'
            'padding:12px 16px;font-size:0.88rem;color:#92400e;">'
            "⚠️ No documents yet. Upload at least one to enable quote generation."
            "</div>"
        )
    total = len(docs)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(0, min(page, total_pages - 1))
    sliced = docs[page * per_page : (page + 1) * per_page]

    count_badge = (
        f'<span style="font-size:0.78rem;color:#64748b;font-weight:400;margin-left:6px;">'
        f'({total} total)</span>'
    )
    rows = "".join(
        f'<tr>'
        f'<td style="padding:8px 12px;border-bottom:1px solid #e8ecf0;color:#374151;">{d["label"]}</td>'
        f'<td style="padding:8px 12px;border-bottom:1px solid #e8ecf0;color:#64748b;font-size:0.82rem;">'
        f'{d["uploaded_at"][:16]}</td>'
        f'</tr>'
        for d in sliced
    )
    return f"""
<div>
  <table style="width:100%;border-collapse:collapse;font-size:0.9rem;
                background:#fff;border-radius:10px;overflow:hidden;
                border:1px solid #e2e8f0;">
    <thead>
      <tr style="background:#f0f4f8;">
        <th style="padding:10px 12px;text-align:left;color:#1e2d45;font-weight:600;
                   border-bottom:1px solid #dde3ec;">Document Type {count_badge}</th>
        <th style="padding:10px 12px;text-align:left;color:#1e2d45;font-weight:600;
                   border-bottom:1px solid #dde3ec;">Uploaded</th>
      </tr>
    </thead>
    <tbody>{rows}</tbody>
  </table>
</div>"""
